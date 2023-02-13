# -*- coding: utf-8 -*-
"""
Created on Sun Feb 12 12:30:36 2023

@author: Mora

9.1 操作EXCEL案例

任务1：
①将wb中的数据写入一个新的wb
②在新的wb中，当delta>0时，标红，delta<0时，标绿
③计算impression>100的url数并写在表格中

"""

import xlrd
import xlwt


def get_style_by_delta(delta):

    style = xlwt.XFStyle()
    font = xlwt.Font()
    font.colour_index
    
    try:   
        if delta < 0:
            font.colour_index = 3
        elif delta > 0:
            font.colour_index = 2
            
    except TypeError:   #when delta = ""，发现类型是str不是float，也就是个空字符串。
        delta = 0
        font.colour_index = 5

    style.font = font
    return style


wb1 = xlrd.open_workbook('D:/富途证券/表格/SEO数据看板1.xls')
sheet1 = wb1.sheet_by_index(0)

data = []  # add all data to this list
for row in range(1, sheet1.nrows):  # every row is a list, add elements to lists
    record = []
    for col in range(sheet1.ncols):
        record.append(sheet1.cell(row, col).value)
    data.append(record)

wb2 = xlwt.Workbook()  # create a wb using object
sheet2 = wb2.add_sheet('带颜色标记的看板')  # :type:sheet
sheet2.write(0, 0, 'Keyword')
sheet2.write(0, 1, 'Impression')
sheet2.write(0, 2, '%Δ')

counter = 0
for row_index, record in enumerate(data):  # start to write data in wb2
    keyword, impression, delta = record    # the structure should be: ['keyword', Impression, delta]
    sheet2.write(row_index + 1, 0, keyword)  # every row_index add 1
    sheet2.write(row_index + 1, 1, impression)
    sheet2.write(row_index + 1, 2, delta, get_style_by_delta(delta))
    if impression > 100:
        counter += 1


sheet2.write(104, 0, '重点关注url数')
sheet2.write(104, 1, counter)

wb2.save('D:/富途证券/表格/SEO数据看板2.xls')


"""
任务2：用openpyxl计算数据，并修改样式
①计算Impression的平均值
②修改字体为等线
③修改单元格的颜色为粉色

"""

import openpyxl
from openpyxl.styles import Font, Alignment

wb = openpyxl.load_workbook('D:/富途证券/表格/SEO数据看板2.xlsx')
sheet= wb.worksheets[0]

font = Font(size=18, bold=True, color='ff1493', name= '等线 Light')
alignment = Alignment(horizontal='center', vertical='center')

sheet['B104'].value = '=average(B2:B102)'     #可以这样直接写入excel公式
sheet['B104'].font = font
sheet['B104'].alignment = alignment

wb.save('D:/富途证券/表格/SEO数据看板2.xlsx')



















