# -*- coding: utf-8 -*-
"""
Created on Wed Jan 25 10:33:32 2023

@author: Mora
持久化操作概述
持久化：将数据从无法长久保存的储存介质转移到可以长久保存的介质中
最简单的方式就是通过 文件系统 将数据保存到 文件中。
"""

'''
【文件操作】
open(filename, mode, encoding)
finally：总是执行代码，不管正常异常，finally中的代码一定会被执行到

一般来说，我们读取文件不会让它全部读完，而是一次性读xx个字符/字节，这样更节省内存空间
'''
#import sys

#print(sys.getdefaultencoding())
import openpyxl
file = open(file= 'C:/Users/Mora/Desktop/致橡树.txt', mode='r', encoding ='utf-8')
try:
    data = file.read(32)
    while data is not None:
        print(data, end = '')
        data = file.read(32)
    #print(file.read(n=16)) #文本文件中，读16个字符
except:
    print('读文件时发生错误！')
finally:
    file.close()#不管正常还是异常，这个文件一定会被关掉，资源会被释放


'''
【读二进制文件】

file seek：文件查找功能
file.seek(0) 将文件指针移动到文件的开头
file.seek(5) 将指针从文件开头向前移动5个字符
file.seek(0,2) 将指针移动到文件末尾
file.seek(5.1) 将指针从当前位置向 前 移动5个字符
file.seek(-5.1)将指针从当前位置向 后 移动5个字符
'''

file = open(file='C:/Users/Mora/Desktop/飞书20221230-145139.jpg', mode='rb') #可以不用给参数明名
#移动文件指针到文件末尾
file.seek(0, 2)
#tell：查看文件指针移动的字节数，这个字节数就是文件的大小
print(file.tell())
file.seek(0)
try:
    data = file.read(512) #磁盘分配空间很多是以512字节为基本的单位
    while data:
        print(data, end='')
        data = file.read(512)
finally:
    file.close()

'''
计算操作MD5的摘要
MD5 Sum 哈希算法，后来SHA-256更常用
文件的MD5码（签名、指纹）
'''

'''
【写文本文件】
w--->覆盖原来的内容，写入新的内容
a--->创建新文件，或者追加。接着之前的内容，写入新的内容，且要记得file.seek(0, 2)将指针异动到末尾。
'''

file = open(file='C:/Users/Mora/Desktop/致橡树写入版.txt', mode='w', encoding ='utf-8')
try:
    file.write('我想做燕子\n')
    file.write('只需要简单思想\n')
    file.write('我想做树\n')

finally:
    file.close()



file = open(file='C:/Users/Mora/Desktop/致橡树写入版.txt', mode='a', encoding ='utf-8')
try:
    file.write('我做不成树\n')
    file.write('因此也撑不破上心的网\n')

finally:
    file.close()

'''
【上下文语法】
with --->进入和离开with的时候会自动执行某些操作。
在离开with上下文的时候，会自动执行file对象的close()方法
'''
with open('C:/Users/Mora/Desktop/致橡树写入版.txt', mode='w',encoding ='utf-8') as file:
    file.write('xxxxx')


'''
【文件的拷贝】
'''
def file_copy(source_file, target_file): #封装一个复制文件的函数
    
    with open(file='C:/Users/Mora/Desktop/飞书20221230-145139.jpg', mode='rb') as source_file:
        with open('C:/Users/Mora/Desktop/核酸检测.jpg', mode='wb') as target_file:
            data = source_file.read(512)
            while data:
                target_file.write(data)
                data = source_file.read(512)
            
#if __name__ == '__main__':
#    file_copy('C:/Users/Mora/Desktop/致橡树.txt','C:/Users/Mora/Desktop/致橡树2.txt')

'''
文件的模式（补充）
x：写入，如果文件已经存在会产生异常
t：文本模式（默认）
+：更新，既可以读又可以写；可以用w+或a+或r+，r+的指针在最开始，w+和a+都在最末尾
'''


'''
【练习1】将100以内的质数输出到文件中，每行一个数
'''



'''
【练习2】
从文件中读入提问测量数据，显示体温不正常的病人的信息
'''


'''
【读写CSV文件】
1. CSV文件--->Comma Seperate Value 逗号分割值文件
    可以是逗号隔开的，也可以是#号、空格隔开的
2. CSB文件的写入：csv.writer(file)  且用gbk编码写入
    writer.writerow([A,B,C])

'''
with open('D:/富途证券/表格/task_output_751 (1).csv', 'r', encoding='utf-8') as file:
    content = file.readline()
    while content:
        values = content.replace('\n','').split(',')
        print(values)
        content = file.readline()
        
        
'''
python里面有一个csv模块，可以用于直接读写csv文件。
'''
import csv

with open('D:/富途证券/表格/task_output_751 (1).csv', 'r', encoding='utf-8') as file:
    reader = csv.reader(file, delimeter='#',quotechar='"') 
    #delimeter-->设置分隔符是#。默认是英文的逗号
    #quotechar-->包围字符的是双引号，但是数据本身并没有双引号，输出的时候希望把双引号去掉。
    for row in reader: #读取每一行的内容
        print(row)

'''
字节序问题：3a和f6谁在上面，谁在下面的顺序
UTF-8 with BOM --->BOM是 字节序标记 byte order marker ---> 带字节序标记（签名）的UFT-8
Pnadas ---> Panle Data Set ---> read_csv / to_csv
'''


'''
【用python进行自动化办公】
1. Python操作Excel
三方库：
~ xlrd / xlwtd / xlutils ---> 兼容低版本的Excel文件(xls)
~ openpyxl ---> Office 2007+ ---> xlsx 

'''

import xlrd

#工作簿是一个Excel文件，即一个workbook
wb = xlrd.open_workbook('D:/富途证券/表格/牛牛圈课堂用户调研统计老.xls')
#获取所有工作表的名字
print(wb.sheet_names())
#获取指定的工作表，worksheet
sheet = wb.sheet_by_name('答卷')
sheet = wb.sheet_by_index(0)
#获取工作表的行 ---> list
print(sheet.nrows, sheet.ncols)
print(sheet.row(2)[2])
#获取指定的列
print(sheet.col(4, start_rowx=1, end_rowx=11))
#获取单元格的数据，value属性可以获取单元格当中的值
cell = sheet.cell(2,2)
print(cell.value) #<C>
#遍历整个excel表单
for row in range(sheet.nrows):
    for col in range(sheet.ncols):
        
        value = sheet.cell(row, col).value
        year, month, date, *_ = xlrd.xldate_as_tuple(value, 0)
        if col == 0:
            print(f'{value:.0f}', end = '\t')
        elif col == 1:
            print(f'{year}年{month:0>2d}月{date:0>2d}日',end='\t')
        else:
            print(value,ent='\t')
    print()    
        

'''
【python的时间日期】

'''
from datetime import datetime

#指定时间日期
date1 = datetime(1990, 5, 3)
print(date1)

#指定年月日时分秒
date2 = datetime(1990, 5, 3, 23, 58, 17)
print(date2)

#现在的时间
date3 = datetime.now()

#计算两个时间相差的时间差对象（timedelta）
delta = date3 - date1

print(delta, type(delta)) #<class 'datetime.timedelta'>
print(delta.days, delta.seconds) #相差多少天、相差多少秒

#格式化时间表达
print(date2.strftime('%Y年%m月%d日 %H时%M分%S秒'))  #月份和日期要小写

'''
【openpyxl 读写】
'''
from datetime import datetime
import openpyxl

wb = openpyxl.load_workbook('D:/富途证券/表格/牛牛圈课堂用户调研统计.xlsx')
print(type(wb))  # <class 'openpyxl.workbook.workbook.Workbook'>

print(wb.sheetnames)  # ['答卷', '统计']

#获取工作表：wb.worksheets
sheet = wb.worksheets[0]
print(type(sheet))  # <class 'openpyxl.worksheet.worksheet.Worksheet'>
print(sheet.dimensions)  # A1:Q413
print(sheet.max_row, sheet.max_column)  # 413 17
print(sheet.rows, sheet.columns)  # 二进制

#遍历所有单元格
print(sheet.cell(1, 1).value)

for row in range(2, sheet.max_row + 1):
    for col in range(1, 14):
        value = sheet.cell(row, col).value
        if col == 2:
            time = datetime.strptime(value,'%Y-%m-%d %H:%M:%S')  #将字符串改为日期格式
            print(time.strftime('%Y年%m月%d日 %H时%M分%S秒'), end='\t') #将日期格式改格式
        elif col == 1:
            print(f'{value:<10d}', end='\t')
        else:
            print(value, end='\t')
    print()

      


#改为更通用的代码：
      
from datetime import datetime
import openpyxl

wb = openpyxl.load_workbook('D:/富途证券/表格/牛牛圈课堂用户调研统计.xlsx')

for row in sheet.rows:
    for cell in row:
        value = cell.value
        if type(value) == datetime.datetime:
            print(value.strftime('%Y年%m月%d日'), end='\t')
        elif type(value) == int:
            print(f'{value:<10d}', end='\t')
        else:
            print(value, end='\t')
        print()


#改为使用单元格的代码：
for row_ch in range(2, 256):
    for col_ch in 'ABCDEF':
        value = sheet[f'{col_ch}{row_ch}'].value
        if type(value) == datetime.datetime:
            print(value.strftime('%Y年%m月%d日'), end='\t')
        elif type(value) == int:
            print(f'{value:<10d}', end='\t')
        else:
            print(value, end='\t')
        print()
        
        
'''
【写Excel文件】
'''
import xlwt

#第一步：创建工作簿(Workbook)
wb = xlwt.Workbook()

#第二步：添加工作表(Worksheet)
sheet = wb.add_sheet('期末成绩') # type: xlwt.Worksheet

#第三步：向单元格写入数据
sheet.write(0, 0, '姓名')
sheet.write(0, 1, '语文')
sheet.write(0, 2, '数字')
sheet.write(0, 3, '语文')

#Todo：写入5个学生3门课程的成绩，成绩用50-100的随机数表示

#第四步：保存工作簿
wb.save('C:/Users/Mora/Desktop/考试成绩记录.xls')












